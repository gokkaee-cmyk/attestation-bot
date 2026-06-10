import asyncio
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _score_color(score):
    if score >= 80:
        return _fill("C6EFCE"), "276221"
    elif score >= 60:
        return _fill("FFEB9C"), "9C6500"
    else:
        return _fill("FFC7CE"), "9C0006"


async def generate_consolidated_report(attestations: list) -> str:
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(None, _build_consolidated, attestations)


def _build_consolidated(attestations: list) -> str:
    wb = Workbook()

    # ── Sheet 1: Итоги по компетенциям ──────────────────────────────────────
    ws1 = wb.active
    ws1.title = "Итоги по компетенциям"

    ws1.merge_cells("A1:G1")
    c = ws1["A1"]
    c.value = f"ИТОГИ АТТЕСТАЦИИ — {datetime.now().strftime('%d.%m.%Y')}"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = _fill("1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    headers = ["ФИО", "Должность", "Компетенция", "%", "Сильные стороны", "Зоны развития", "Рекомендации"]
    col_widths = [30, 28, 26, 8, 40, 40, 45]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws1.cell(row=2, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = _fill("2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws1.column_dimensions[get_column_letter(col_idx)].width = w
    ws1.row_dimensions[2].height = 28
    ws1.freeze_panes = "A3"

    row = 3
    for att in attestations:
        name = att.get("name", "")
        position = att.get("position_name", "")
        competency_avg = att.get("competency_avg", {})
        answers = att.get("answers", [])

        by_comp = {}
        for a in answers:
            comp = a.get("competency", "")
            if comp not in by_comp:
                by_comp[comp] = []
            by_comp[comp].append(a)

        # Person header row
        ws1.merge_cells(f"A{row}:G{row}")
        header_cell = ws1.cell(row=row, column=1, value=f"{name}  |  {position}")
        header_cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        header_cell.fill = _fill("2E4057")
        header_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        header_cell.border = _border()
        ws1.row_dimensions[row].height = 22
        row += 1

        first_person = True
        for comp, avg in competency_avg.items():
            avg_pct = round(avg)
            comp_answers = by_comp.get(comp, [])
            strengths = " | ".join([a.get("strengths", "") for a in comp_answers if a.get("strengths")])
            weaknesses = " | ".join([a.get("weaknesses", "") for a in comp_answers if a.get("weaknesses")])
            recommendations = " | ".join([a.get("recommendation", "") for a in comp_answers if a.get("recommendation")])

            row_data = [
                name if first_person else "",
                position if first_person else "",
                comp,
                f"{avg_pct}%",
                strengths,
                weaknesses,
                recommendations,
            ]

            score_fill, score_font_color = _score_color(avg_pct)

            for col_idx, val in enumerate(row_data, start=1):
                cell = ws1.cell(row=row, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.fill = _fill("EBF3FB")
                cell.alignment = Alignment(
                    horizontal="center" if col_idx == 4 else "left",
                    vertical="center", wrap_text=True,
                    indent=0 if col_idx == 4 else 1
                )
                cell.border = _border()
                if col_idx == 4:
                    cell.fill = score_fill
                    cell.font = Font(name="Arial", size=10, bold=True, color=score_font_color)

            ws1.row_dimensions[row].height = 40
            row += 1
            first_person = False

        # Average row
        overall = round(att.get("overall_avg", 0))
        avg_fill, avg_font = _score_color(overall)
        for col_idx in range(1, 8):
            cell = ws1.cell(row=row, column=col_idx, value="")
            cell.fill = _fill("D9E1F2")
            cell.border = _border()
        ws1.cell(row=row, column=1, value=name).font = Font(name="Arial", size=10, bold=True)
        ws1.cell(row=row, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws1.cell(row=row, column=3, value="Средний % по всем компетенциям").font = Font(name="Arial", size=10, bold=True)
        ws1.cell(row=row, column=3).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        avg_cell = ws1.cell(row=row, column=4, value=f"{overall}%")
        avg_cell.font = Font(name="Arial", size=10, bold=True, color=avg_font)
        avg_cell.fill = avg_fill
        avg_cell.alignment = Alignment(horizontal="center", vertical="center")
        ws1.row_dimensions[row].height = 20
        row += 1

        # Spacer
        for col_idx in range(1, 8):
            cell = ws1.cell(row=row, column=col_idx, value="")
            cell.fill = _fill("F2F2F2")
            cell.border = _border()
        ws1.row_dimensions[row].height = 6
        row += 1

    # ── Sheet 2: Сводная таблица ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Сводная таблица")

    ws2.merge_cells("A1:E1")
    c2 = ws2["A1"]
    c2.value = f"СВОДНАЯ ТАБЛИЦА АТТЕСТАЦИИ — {datetime.now().strftime('%d.%m.%Y')}"
    c2.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c2.fill = _fill("1F3864")
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    headers2 = ["ФИО", "Должность", "Средний %", "Статус аттестации", "Общий комментарий"]
    col_widths2 = [32, 30, 14, 30, 60]
    for col_idx, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = _fill("2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[2].height = 28
    ws2.freeze_panes = "A3"

    def level(s):
        if s >= 90: return "отлично"
        if s >= 80: return "хорошо"
        if s >= 60: return "средний уровень"
        return "требует развития"

    for row_idx, att in enumerate(attestations, start=3):
        name = att.get("name", "")
        position = att.get("position_name", "")
        overall = round(att.get("overall_avg", 0))
        verdict = att.get("verdict", "")
        competency_avg = att.get("competency_avg", {})

        comment = " | ".join(
            f"{comp}: {round(avg)}% — {level(round(avg))}"
            for comp, avg in competency_avg.items()
        )

        row_data = [name, position, f"{overall}%", verdict, comment]
        score_fill, score_font_color = _score_color(overall)
        fill_color = "EBF3FB" if row_idx % 2 == 0 else "FFFFFF"

        for col_idx, val in enumerate(row_data, start=1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.fill = _fill(fill_color)
            cell.alignment = Alignment(
                horizontal="center" if col_idx in (3, 4) else "left",
                vertical="center", wrap_text=True,
                indent=0 if col_idx in (3, 4) else 1
            )
            cell.border = _border()
            if col_idx in (3, 4):
                cell.fill = score_fill
                cell.font = Font(name="Arial", size=10, bold=True, color=score_font_color)

        ws2.row_dimensions[row_idx].height = 30

    output_path = f"/tmp/consolidated_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_path)
    return output_path
