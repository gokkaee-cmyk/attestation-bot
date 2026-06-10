import asyncio
from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _border(style="thin"):
    s = Side(style=style)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


async def generate_report(name, position_name, position_key, answers,
                           competency_avg, overall_avg, verdict, start_time):
    loop = asyncio.get_event_loop()
    return await loop.run_in_executor(
        None, _build_excel, name, position_name, answers,
        competency_avg, overall_avg, verdict, start_time
    )


def _build_excel(name, position_name, answers, competency_avg, overall_avg, verdict, start_time):
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Итоги по компетенциям"

    ws1.merge_cells("A1:F1")
    c = ws1["A1"]
    c.value = f"Итоги аттестации — {name} — {position_name}"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = _fill("1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[1].height = 32

    ws1.merge_cells("A2:F2")
    d = ws1["A2"]
    d.value = f"Дата аттестации: {datetime.fromisoformat(start_time).strftime('%d.%m.%Y %H:%M')}"
    d.font = Font(name="Arial", size=10, italic=True)
    d.fill = _fill("D9E1F2")
    d.alignment = Alignment(horizontal="center", vertical="center")
    ws1.row_dimensions[2].height = 18

    headers = ["ФИО", "Компетенция", "%", "Сильные стороны", "Зоны развития", "Рекомендации"]
    col_widths = [28, 26, 8, 38, 38, 42]
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws1.cell(row=3, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = _fill("2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws1.column_dimensions[get_column_letter(col_idx)].width = w
    ws1.row_dimensions[3].height = 28
    ws1.freeze_panes = "A4"

    comp_rows = {}
    for a in answers:
        comp = a["competency"]
        if comp not in comp_rows:
            comp_rows[comp] = []
        comp_rows[comp].append(a)

    row = 4
    alt = False
    for comp, items in comp_rows.items():
        fill_color = "EBF3FB" if not alt else "FFFFFF"
        alt = not alt
        first = True
        for a in items:
            score_val = round(a["score"])
            row_data = [
                name if first else "",
                a["competency"] if first else "",
                score_val,
                a.get("strengths", ""),
                a.get("weaknesses", ""),
                a["recommendation"],
            ]
            for col_idx, val in enumerate(row_data, start=1):
                cell = ws1.cell(row=row, column=col_idx, value=val)
                cell.font = Font(name="Arial", size=10)
                cell.fill = _fill(fill_color)
                cell.alignment = Alignment(
                    horizontal="center" if col_idx in (1, 2, 3) else "left",
                    vertical="top", wrap_text=True,
                    indent=0 if col_idx in (1, 2, 3) else 1
                )
                cell.border = _border()
                if col_idx == 3:
                    if score_val >= 80:
                        cell.fill = _fill("C6EFCE")
                        cell.font = Font(name="Arial", size=10, bold=True, color="276221")
                    elif score_val >= 60:
                        cell.fill = _fill("FFEB9C")
                        cell.font = Font(name="Arial", size=10, bold=True, color="9C6500")
                    else:
                        cell.fill = _fill("FFC7CE")
                        cell.font = Font(name="Arial", size=10, bold=True, color="9C0006")
            ws1.row_dimensions[row].height = 60
            first = False
            row += 1

        avg_val = round(competency_avg.get(comp, 0))
        for col_idx in range(1, 7):
            ws1.cell(row=row, column=col_idx).fill = _fill("D9E1F2")
            ws1.cell(row=row, column=col_idx).border = _border()
            ws1.cell(row=row, column=col_idx).alignment = Alignment(horizontal="center", vertical="center")
        ws1.cell(row=row, column=1, value=name).font = Font(name="Arial", size=10, bold=True)
        ws1.cell(row=row, column=2, value=comp).font = Font(name="Arial", size=10, bold=True)
        avg_cell = ws1.cell(row=row, column=3, value=avg_val)
        ws1.cell(row=row, column=4, value="Средний % по компетенции").font = Font(name="Arial", size=10, italic=True)
        if avg_val >= 80:
            avg_cell.fill = _fill("C6EFCE")
            avg_cell.font = Font(name="Arial", size=10, bold=True, color="276221")
        elif avg_val >= 60:
            avg_cell.fill = _fill("FFEB9C")
            avg_cell.font = Font(name="Arial", size=10, bold=True, color="9C6500")
        else:
            avg_cell.fill = _fill("FFC7CE")
            avg_cell.font = Font(name="Arial", size=10, bold=True, color="9C0006")
        ws1.row_dimensions[row].height = 20
        row += 1

    ws2 = wb.create_sheet("Сводная таблица")

    ws2.merge_cells("A1:D1")
    c = ws2["A1"]
    c.value = "СВОДНАЯ ТАБЛИЦА АТТЕСТАЦИИ"
    c.font = Font(name="Arial", bold=True, size=13, color="FFFFFF")
    c.fill = _fill("1F3864")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws2.row_dimensions[1].height = 32

    headers2 = ["ФИО", "Средний %", "Статус аттестации", "Общий комментарий"]
    col_widths2 = [32, 14, 30, 60]
    for col_idx, (h, w) in enumerate(zip(headers2, col_widths2), start=1):
        cell = ws2.cell(row=2, column=col_idx, value=h)
        cell.font = Font(name="Arial", bold=True, size=11, color="FFFFFF")
        cell.fill = _fill("2E75B6")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border()
        ws2.column_dimensions[get_column_letter(col_idx)].width = w
    ws2.row_dimensions[2].height = 28

    overall_val = round(overall_avg)
    comp_comments = []
    for comp, avg in competency_avg.items():
        avg_r = round(avg)
        if avg_r >= 80:
            comp_comments.append(f"{comp}: {avg_r}% ✓")
        else:
            comp_comments.append(f"{comp}: {avg_r}% — требует развития")
    general_comment = " | ".join(comp_comments)

    row_data = [name, f"{overall_val}%", verdict, general_comment]
    for col_idx, val in enumerate(row_data, start=1):
        cell = ws2.cell(row=3, column=col_idx, value=val)
        cell.font = Font(name="Arial", size=11)
        cell.alignment = Alignment(
            horizontal="center" if col_idx in (1, 2, 3) else "left",
            vertical="center", wrap_text=True,
            indent=0 if col_idx in (1, 2, 3) else 1
        )
        cell.border = _border()
        if col_idx in (2, 3):
            if overall_val >= 80:
                cell.fill = _fill("C6EFCE")
                cell.font = Font(name="Arial", size=11, bold=True, color="276221")
            else:
                cell.fill = _fill("FFC7CE")
                cell.font = Font(name="Arial", size=11, bold=True, color="9C0006")
        else:
            cell.fill = _fill("EBF3FB")
    ws2.row_dimensions[3].height = 50

    output_path = f"/tmp/attestation_{name.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    wb.save(output_path)
    return output_path
